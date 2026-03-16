import tkinter as tk
from tkinter import filedialog,messagebox
import pandas as pd
import random
import threading
import time
from datetime import datetime
from openpyxl import Workbook,load_workbook
from fpdf import FPDF
import os
import sys
import pystray
from PIL import Image
import json

ROWS=5
COLS=6
TOTAL=30

seat_names=[]
seats=[]
locked=set()

duty_students=[]
duty_queue=[]

seat_file="seat_layout.xlsx"
seat_names_file="seat_names.xlsx"
duty_names_file="duty_names.xlsx"
history_file="Duty_History.xlsx"
fixed_file="fixed_duty.json"

tasks=[
"擦黑板",
"教室扫拖",
"倒垃圾",
"室外卫生区",
"组长，负责各种擦(doge)"
]

fixed_duty={t:None for t in tasks}

drag_data={"row":None,"col":None}

# ----------------------
# 读取固定值日
# ----------------------

def load_fixed_duty():
    global fixed_duty
    if os.path.exists(fixed_file):
        with open(fixed_file,"r",encoding="utf-8") as f:
            fixed_duty=json.load(f)

def save_fixed_duty():
    with open(fixed_file,"w",encoding="utf-8") as f:
        json.dump(fixed_duty,f,ensure_ascii=False,indent=2)

# ----------------------
# 名单加载
# ----------------------

def load_saved_seat_names():
    global seat_names
    if os.path.exists(seat_names_file):
        df=pd.read_excel(seat_names_file,header=None)
        seat_names=df.iloc[:,0].dropna().tolist()

def load_saved_duty_names():
    global duty_students
    if os.path.exists(duty_names_file):
        df=pd.read_excel(duty_names_file,header=None)
        duty_students=df.iloc[:,0].dropna().tolist()

# ----------------------
# 导入名单
# ----------------------

def load_seat_excel():
    global seat_names
    file=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls")])
    if not file:return
    df=pd.read_excel(file,header=None)
    seat_names=df.iloc[:,0].dropna().tolist()
    pd.DataFrame(seat_names).to_excel(seat_names_file,index=False,header=False)
    messagebox.showinfo("成功",f"读取 {len(seat_names)} 名学生")

def load_duty_excel():
    global duty_students
    file=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls")])
    if not file:return
    df=pd.read_excel(file,header=None)
    duty_students=df.iloc[:,0].dropna().tolist()
    pd.DataFrame(duty_students).to_excel(duty_names_file,index=False,header=False)
    messagebox.showinfo("成功",f"读取 {len(duty_students)} 名学生")

# ----------------------
# 座位
# ----------------------

def generate_seats():

    global seats

    if not seat_names:
        messagebox.showwarning("提示","请先导入名单")
        return

    shuffled=seat_names[:]
    random.shuffle(shuffled)

    while len(shuffled)<TOTAL:
        shuffled.append("")

    seats=[shuffled[i*COLS:(i+1)*COLS] for i in range(ROWS)]

    draw_seats()
    save_seat_layout()

def save_seat_layout():

    wb=Workbook()
    ws=wb.active

    for r in range(ROWS):
        for c in range(COLS):
            ws.cell(r+1,c+1).value=seats[r][c]

    wb.save(seat_file)

def load_seat_layout():

    global seats

    if os.path.exists(seat_file):

        wb=load_workbook(seat_file)
        ws=wb.active

        seats=[]

        for r in range(ROWS):

            row=[]

            for c in range(COLS):

                val=ws.cell(r+1,c+1).value
                row.append(val if val else "")

            seats.append(row)

    else:

        seats=[[""]*COLS for _ in range(ROWS)]

# ----------------------
# 拖拽座位
# ----------------------

def on_click(event):

    x=event.x
    y=event.y

    cell_w=110
    cell_h=50

    col=(x-40)//cell_w
    row=(y-100)//cell_h

    if 0<=row<ROWS and 0<=col<COLS:
        drag_data["row"]=row
        drag_data["col"]=col

def on_release(event):

    cell_w=110
    cell_h=50

    col=(event.x-40)//cell_w
    row=(event.y-100)//cell_h

    r1=drag_data["row"]
    c1=drag_data["col"]

    if r1 is None:return

    if 0<=row<ROWS and 0<=col<COLS:

        if (r1,c1) in locked or (row,col) in locked:
            messagebox.showwarning("提示","座位已锁定")
            return

        seats[r1][c1],seats[row][col]=seats[row][col],seats[r1][c1]

    drag_data["row"]=None

    draw_seats()
    save_seat_layout()

# ----------------------
# 绘制座位
# ----------------------

def draw_seats():

    canvas.delete("all")

    cell_w=110
    cell_h=50

    start_x=40
    start_y=100

    canvas.create_rectangle(start_x,30,start_x+COLS*cell_w,70,fill="#ddd")
    canvas.create_text(start_x+COLS*cell_w/2,50,text="讲台",font=("Arial",14,"bold"))

    for r in range(ROWS):
        for c in range(COLS):

            x1=start_x+c*cell_w
            y1=start_y+r*cell_h

            x2=x1+cell_w
            y2=y1+cell_h

            color="#fff"

            if (r,c) in locked:
                color="#ffd6d6"

            canvas.create_rectangle(x1,y1,x2,y2,fill=color)

            name=seats[r][c]

            canvas.create_text((x1+x2)/2,(y1+y2)/2,text=name,font=("Arial",10))

# ----------------------
# 锁定座位
# ----------------------

def lock_seat():

    if drag_data["row"] is None:return

    r=drag_data["row"]
    c=drag_data["col"]

    if (r,c) in locked:
        locked.remove((r,c))
    else:
        locked.add((r,c))

    draw_seats()

# ----------------------
# 固定值日设置
# ----------------------

def set_fixed_duty():

    if not duty_students:
        messagebox.showwarning("提示","请先导入值日名单")
        return

    win=tk.Toplevel()
    win.title("固定值日设置")

    vars={}

    for i,task in enumerate(tasks):

        tk.Label(win,text=task,width=18,anchor="w").grid(row=i,column=0,padx=5,pady=5)

        var=tk.StringVar()
        var.set(fixed_duty.get(task) or "随机")

        vars[task]=var

        options=["随机"]+duty_students

        menu=tk.OptionMenu(win,var,*options)
        menu.grid(row=i,column=1,padx=5,pady=5)

    def save():

        for t in tasks:

            v=vars[t].get()

            if v=="随机":
                fixed_duty[t]=None
            else:
                fixed_duty[t]=v

        save_fixed_duty()

        messagebox.showinfo("成功","固定值日已保存")

        win.destroy()

    tk.Button(win,text="保存",command=save).grid(row=len(tasks),columnspan=2,pady=10)

# ----------------------
# 值日生成
# ----------------------

def get_today_duty():

    global duty_queue

    result=[None]*5
    used=set()

    for i,task in enumerate(tasks):

        person=fixed_duty.get(task)

        if person:
            result[i]=person
            used.add(person)

    for i in range(5):

        if result[i] is None:

            while True:

                if len(duty_queue)==0:
                    duty_queue=duty_students[:]
                    random.shuffle(duty_queue)

                s=duty_queue.pop(0)

                if s not in used:
                    result[i]=s
                    used.add(s)
                    break

    return result

# ----------------------
# 显示值日
# ----------------------

def show_duty_window():

    today=get_today_duty()

    roles=["甲","乙","丙","丁","戊"]

    text=""

    for r,s,t in zip(roles,today,tasks):
        text+=f"{r}：{s} {t}\n"

    win=tk.Toplevel()
    win.title("今日值日")
    win.geometry("380x200")

    tk.Label(win,text=text,font=("Arial",14)).pack(pady=30)

    save_history(today)
    save_pdf(today)

# ----------------------
# 一周值日
# ----------------------

def weekly_duty():

    text=""

    for i in range(5):

        today=get_today_duty()

        roles=["甲","乙","丙","丁","戊"]

        text+=f"第{i+1}天\n"

        for r,s,t in zip(roles,today,tasks):
            text+=f"{r}：{s} {t}\n"

        text+="\n"

    win=tk.Toplevel()
    win.title("一周值日表")
    tk.Label(win,text=text,font=("Arial",12)).pack()

# ----------------------
# 值日统计
# ----------------------

def duty_stats():

    if not os.path.exists(history_file):
        messagebox.showinfo("提示","暂无记录")
        return

    wb=load_workbook(history_file)
    ws=wb.active

    stats={}

    for row in ws.iter_rows(min_row=2,values_only=True):

        for name in row[1:]:

            stats[name]=stats.get(name,0)+1

    text="值日统计\n\n"

    for k,v in stats.items():
        text+=f"{k} : {v}\n"

    win=tk.Toplevel()
    win.title("值日统计")
    tk.Label(win,text=text,font=("Arial",12)).pack()

# ----------------------
# 保存记录
# ----------------------

def save_history(today):

    date=datetime.now().strftime("%Y-%m-%d")

    if os.path.exists(history_file):

        wb=load_workbook(history_file)
        ws=wb.active

    else:

        wb=Workbook()
        ws=wb.active
        ws.append(["日期","甲","乙","丙","丁","戊"])

    if ws.max_row>1 and ws.cell(ws.max_row,1).value==date:
        return

    ws.append([date]+today)

    wb.save(history_file)

def save_pdf(today):

    pdf=FPDF()
    pdf.add_page()

    pdf.set_font("Arial","B",16)

    pdf.cell(0,10,"今日值日",ln=1,align="C")

    roles=["甲","乙","丙","丁","戊"]

    pdf.set_font("Arial","",14)

    for r,s,t in zip(roles,today,tasks):
        pdf.cell(0,10,f"{r}：{s} {t}",ln=1)

    pdf.output("Duty_"+datetime.now().strftime("%Y%m%d")+".pdf")

# ----------------------
# 自动时间
# ----------------------

def duty_timer():

    while True:

        now=time.localtime()

        if now.tm_hour==20 and now.tm_min==40:

            root.after(0,show_duty_window)

            time.sleep(60)

        time.sleep(20)

# ----------------------
# 托盘
# ----------------------

def tray_show(icon,item):
    root.deiconify()

def tray_quit(icon,item):
    icon.stop()
    root.quit()

def start_tray():

    image=Image.open("icon.ico")

    icon=pystray.Icon(
        "SmartClass",
        image,
        "智慧班级",
        menu=pystray.Menu(
            pystray.MenuItem("显示主窗口",tray_show),
            pystray.MenuItem("退出",tray_quit)
        )
    )

    icon.run()

# ----------------------
# GUI
# ----------------------

root=tk.Tk()
root.title("智慧班级 v3")
root.geometry("860x580")

load_saved_seat_names()
load_saved_duty_names()
load_seat_layout()
load_fixed_duty()

if len(sys.argv)>1 and sys.argv[1]=="--startup":
    root.withdraw()
    threading.Thread(target=start_tray,daemon=True).start()

top=tk.Frame(root)
top.pack(pady=10)

tk.Button(top,text="导入座位名单",command=load_seat_excel).grid(row=0,column=0,padx=5)
tk.Button(top,text="随机排座",command=generate_seats).grid(row=0,column=1,padx=5)
tk.Button(top,text="锁定座位",command=lock_seat).grid(row=0,column=2,padx=5)

tk.Button(top,text="导入值日名单",command=load_duty_excel).grid(row=0,column=3,padx=5)
tk.Button(top,text="生成今日值日",command=show_duty_window).grid(row=0,column=4,padx=5)
tk.Button(top,text="一周值日表",command=weekly_duty).grid(row=0,column=5,padx=5)
tk.Button(top,text="值日统计",command=duty_stats).grid(row=0,column=6,padx=5)
tk.Button(top,text="固定值日设置",command=set_fixed_duty).grid(row=0,column=7,padx=5)

canvas=tk.Canvas(root,width=800,height=420)
canvas.pack()

canvas.bind("<ButtonPress-1>",on_click)
canvas.bind("<ButtonRelease-1>",on_release)

draw_seats()

threading.Thread(target=duty_timer,daemon=True).start()

root.mainloop()
